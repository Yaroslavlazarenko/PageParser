import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Dict

def save_results_to_excel(file_path: str, students_data: List[Dict], col_name: str):
    """
    Відкриває Excel-файл і записує результати, орієнтуючись на позицію стовпця "Телефон".
    """
    print(f"\n💾 Зберігаю всі результати у файл '{os.path.basename(file_path)}'...")
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        header = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        col_index = None

        # --- НОВА, БІЛЬШ НАДІЙНА ЛОГІКА ПОШУКУ СТОВПЦЯ ---
        try:
            # 1. Шукаємо стовпець "Телефон" як надійний орієнтир
            phone_col_index = header.index("телефон") + 1
            # Наш цільовий стовпець знаходиться через один після "Телефону"
            col_index = phone_col_index + 2
            
            # Перевіримо, чи є у цього стовпця заголовок. Якщо ні, додамо.
            if sheet.cell(row=1, column=col_index).value is None:
                sheet.cell(row=1, column=col_index, value=col_name)

        except ValueError:
            # 2. Якщо "Телефон" не знайдено, повертаємося до старої логіки
            print("⚠️ Стовпець 'Телефон' не знайдено. Спробую знайти за назвою '{col_name}'.")
            try:
                col_index = header.index(col_name.strip().lower()) + 1
            except ValueError:
                print(f"⚠️ Стовпець '{col_name}' також не знайдено. Створюю новий в кінці таблиці.")
                col_index = sheet.max_column + 1
                sheet.cell(row=1, column=col_index, value=col_name)
        # --- КІНЕЦЬ НОВОЇ ЛОГІКИ ---

        # Оновлюємо комірки для кожного студента з результатом
        for student in students_data:
            if 'final_result' in student:
                excel_row = student['index'] + 2
                sheet.cell(row=excel_row, column=col_index, value=student['final_result'])
        
        workbook.save(filename=file_path)
        print("✅ Файл успішно збережено!")

    except (PermissionError, IOError):
        print(f"❌ ПОМИЛКА ЗАПИСУ: Не вдалося зберегти файл. "
              f"Переконайтесь, що він не відкритий в іншій програмі.")
    except (InvalidFileException, FileNotFoundError):
        print(f"❌ ПОМИЛКА: Не вдалося відкрити файл. Можливо, він пошкоджений.")
    except Exception as e:
        print(f"❌ НЕОЧІКУВАНА ПОМИЛКА під час запису в Excel: {e}")